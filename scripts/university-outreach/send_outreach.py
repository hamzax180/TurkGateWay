#!/usr/bin/env python3
"""
Outreach mailer for the Turkish foundation-university list.

Sends one personalised message per university from international@turkgateway.com,
reading recipients straight out of universities/turkey_private_university_emails.xlsx.

NOTHING IS SENT unless you pass BOTH --send and --confirm. Every other mode is a
dry run that only prints what would go out.

    # see what would be sent (default; sends nothing)
    python scripts/university-outreach/send_outreach.py

    # read three fully rendered emails end to end
    python scripts/university-outreach/send_outreach.py --preview 3

    # one real email, to yourself, to check formatting and deliverability
    python scripts/university-outreach/send_outreach.py --test-to you@turkgateway.com --send --confirm

    # the real run
    python scripts/university-outreach/send_outreach.py --send --confirm

Credentials come from the environment (or .env.local) and are never written to
disk or printed by this script. See README.md in this folder.
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import smtplib
import ssl
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from email.headerregistry import Address
from email.message import EmailMessage
from email.utils import formatdate, make_msgid
from pathlib import Path

try:
    import openpyxl
except ImportError:  # pragma: no cover - environment guard
    sys.exit("openpyxl is required:  python -m pip install openpyxl")


HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
DEFAULT_WORKBOOK = REPO / "universities" / "turkey_private_university_emails.xlsx"
DEFAULT_SHEET = "Private Universities - Turkey"
SENT_LOG = HERE / "sent_log.json"
TEMPLATE_TXT = HERE / "template.txt"
TEMPLATE_HTML = HERE / "template.html"

# Placeholders the template ships with. Sending any of these to a real
# university would be worse than not sending at all, so the run is refused
# while one survives into the rendered message.
UNFILLED_RE = re.compile(r"\[[A-Z][A-Z0-9 _/&-]{2,}\]")


# ── recipients ───────────────────────────────────────────────────────────────

@dataclass
class Recipient:
    university: str
    city: str
    website: str
    email: str
    department: str
    status: str
    notes: str

    @property
    def key(self) -> str:
        return self.email.strip().lower()

    @property
    def greeting(self) -> str:
        """'Dear International Admissions team,' beats 'Dear Sir or Madam,'."""
        dept = (self.department or "").strip()
        if dept and not dept.lower().startswith("general"):
            return f"{dept} team"
        return "International Office"

    @property
    def domain(self) -> str:
        return self.key.split("@")[-1] if "@" in self.key else ""


def load_recipients(workbook: Path, sheet: str) -> list[Recipient]:
    if not workbook.exists():
        sys.exit(f"Workbook not found: {workbook}")

    wb = openpyxl.load_workbook(workbook, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet {sheet!r} not in {workbook.name}. Found: {wb.sheetnames}")
    ws = wb[sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        sys.exit("Sheet is empty.")

    header = [str(c or "").strip().lower() for c in rows[0]]

    def col(*names: str) -> int | None:
        for name in names:
            if name in header:
                return header.index(name)
        return None

    idx = {
        "university": col("university", "name"),
        "city": col("city"),
        "website": col("website", "url"),
        "email": col("best contact email", "contact email", "email"),
        "department": col("department / office", "department", "office"),
        "status": col("status"),
        "notes": col("notes"),
    }
    if idx["university"] is None or idx["email"] is None:
        sys.exit(f"Could not find University/Email columns. Header was: {header}")

    def cell(row, key: str) -> str:
        i = idx[key]
        if i is None or i >= len(row):
            return ""
        return str(row[i] or "").strip()

    out: list[Recipient] = []
    seen: set[str] = set()
    for row in rows[1:]:
        if not row or not any(row):
            continue
        email = cell(row, "email")
        # A malformed address aborts the whole run rather than being quietly
        # skipped — a typo here means a university silently never hears from us.
        if not email:
            continue
        if "@" not in email or " " in email:
            sys.exit(f"Malformed address for {cell(row, 'university')!r}: {email!r}")
        r = Recipient(
            university=cell(row, "university"),
            city=cell(row, "city"),
            website=cell(row, "website"),
            email=email,
            department=cell(row, "department"),
            status=cell(row, "status"),
            notes=cell(row, "notes"),
        )
        if r.key in seen:
            continue
        seen.add(r.key)
        out.append(r)
    return out


# ── template ─────────────────────────────────────────────────────────────────

def render(text: str, r: Recipient, extra: dict[str, str]) -> str:
    values = {
        "university": r.university,
        "city": r.city,
        "website": r.website,
        "email": r.email,
        "department": r.department,
        "greeting": r.greeting,
        **extra,
    }
    for key, value in values.items():
        text = text.replace("{{" + key + "}}", value)
    return text


def load_template() -> tuple[str, str, str | None]:
    """Returns (subject_template, text_template, html_template|None)."""
    if not TEMPLATE_TXT.exists():
        sys.exit(f"Missing template: {TEMPLATE_TXT}")

    raw = TEMPLATE_TXT.read_text(encoding="utf-8")
    lines = raw.splitlines()
    if not lines or not lines[0].lower().startswith("subject:"):
        sys.exit("template.txt must start with a 'Subject: ...' line.")

    subject = lines[0].split(":", 1)[1].strip()
    body = "\n".join(lines[1:]).lstrip("\n")

    html = TEMPLATE_HTML.read_text(encoding="utf-8") if TEMPLATE_HTML.exists() else None
    return subject, body, html


# ── send log ─────────────────────────────────────────────────────────────────

def load_log() -> dict:
    if not SENT_LOG.exists():
        return {}
    try:
        return json.loads(SENT_LOG.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        sys.exit(f"{SENT_LOG} is corrupt. Fix or delete it before running again.")


def save_log(log: dict) -> None:
    SENT_LOG.write_text(json.dumps(log, indent=2, ensure_ascii=False), encoding="utf-8")


# ── config ───────────────────────────────────────────────────────────────────

def load_dotenv(path: Path) -> None:
    """Fills os.environ from .env.local without overriding a real env var."""
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key, value = key.strip(), value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


@dataclass
class SmtpConfig:
    host: str
    port: int
    user: str
    password: str
    from_name: str
    from_addr: str
    reply_to: str
    use_ssl: bool

    @classmethod
    def from_env(cls) -> "SmtpConfig":
        from_addr = os.environ.get("OUTREACH_FROM", "international@turkgateway.com")
        port = int(os.environ.get("SMTP_PORT", "587"))
        return cls(
            host=os.environ.get("SMTP_HOST", ""),
            port=port,
            user=os.environ.get("SMTP_USER", from_addr),
            password=os.environ.get("SMTP_PASSWORD", ""),
            from_name=os.environ.get("OUTREACH_FROM_NAME", "TurkGateway"),
            from_addr=from_addr,
            reply_to=os.environ.get("OUTREACH_REPLY_TO", from_addr),
            use_ssl=port == 465,
        )

    def require(self) -> None:
        missing = [n for n, v in (("SMTP_HOST", self.host), ("SMTP_PASSWORD", self.password)) if not v]
        if missing:
            sys.exit(
                "Missing required environment variable(s): "
                + ", ".join(missing)
                + "\nSet them in your shell or .env.local — see scripts/university-outreach/README.md."
            )


# ── message building ─────────────────────────────────────────────────────────

def build_message(cfg: SmtpConfig, r: Recipient, subject: str, text: str,
                  html: str | None, to_override: str | None) -> EmailMessage:
    msg = EmailMessage()
    local, _, domain = cfg.from_addr.partition("@")
    msg["From"] = Address(cfg.from_name, local, domain)
    msg["To"] = to_override or r.email
    msg["Subject"] = subject
    msg["Reply-To"] = cfg.reply_to
    msg["Date"] = formatdate(localtime=True)
    msg["Message-ID"] = make_msgid(domain=domain or "turkgateway.com")
    # One-to-one business correspondence, but an easy opt-out is still the right
    # thing to offer and keeps the sending domain in good standing.
    msg["List-Unsubscribe"] = f"<mailto:{cfg.reply_to}?subject=unsubscribe>"
    msg["Auto-Submitted"] = "auto-generated"

    msg.set_content(text, subtype="plain", charset="utf-8")
    if html:
        msg.add_alternative(html, subtype="html")
    return msg


def check_unfilled(subject: str, text: str, html: str | None, who: str) -> None:
    blob = "\n".join(p for p in (subject, text, html or "") if p)
    hits = sorted(set(UNFILLED_RE.findall(blob)))
    if hits:
        sys.exit(
            f"Template still contains unfilled placeholders (seen while rendering {who}):\n"
            + "\n".join(f"  {h}" for h in hits)
            + "\n\nEdit scripts/university-outreach/template.txt (and template.html) first."
        )


# ── main ─────────────────────────────────────────────────────────────────────

def main() -> None:
    ap = argparse.ArgumentParser(
        description="Personalised outreach mailer for Turkish foundation universities.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    ap.add_argument("--send", action="store_true",
                    help="actually transmit. Without this the run is a dry run.")
    ap.add_argument("--confirm", action="store_true",
                    help="required alongside --send. Two flags, so no single typo sends 43 emails.")
    ap.add_argument("--status", choices=["confirmed", "inferred", "all"], default="confirmed",
                    help="which addresses to include (default: confirmed only — the "
                         "'Inferred (pattern)' rows are guesses and are likely to bounce).")
    ap.add_argument("--limit", type=int, default=0, help="stop after N recipients (0 = no limit).")
    ap.add_argument("--delay", type=float, default=20.0,
                    help="seconds to wait between sends, plus jitter (default: 20).")
    ap.add_argument("--test-to", metavar="ADDR",
                    help="redirect every message to ADDR instead of the university.")
    ap.add_argument("--preview", type=int, metavar="N", default=0,
                    help="print N fully rendered emails and exit.")
    ap.add_argument("--only", metavar="SUBSTR",
                    help="only universities whose name or email contains SUBSTR.")
    ap.add_argument("--resend", action="store_true",
                    help="ignore sent_log.json and include already-contacted addresses.")
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--sheet", default=DEFAULT_SHEET)
    args = ap.parse_args()

    load_dotenv(REPO / ".env.local")
    cfg = SmtpConfig.from_env()

    subject_tpl, text_tpl, html_tpl = load_template()
    recipients = load_recipients(args.workbook, args.sheet)
    log = {} if args.resend else load_log()

    # filters
    if args.status != "all":
        want = "confirmed" if args.status == "confirmed" else "inferred"
        recipients = [r for r in recipients if want in (r.status or "").lower()]
    if args.only:
        needle = args.only.lower()
        recipients = [r for r in recipients
                      if needle in r.university.lower() or needle in r.key]

    skipped = [r for r in recipients if r.key in log]
    queue = [r for r in recipients if r.key not in log]
    if args.limit:
        queue = queue[: args.limit]

    extra = {"intake": os.environ.get("OUTREACH_INTAKE", "2026/2027")}

    print(f"Workbook   : {args.workbook.name}  ({args.sheet})")
    print(f"From       : {cfg.from_name} <{cfg.from_addr}>")
    print(f"SMTP       : {cfg.host or '(unset)'}:{cfg.port}"
          f"{'  [SSL]' if cfg.use_ssl else '  [STARTTLS]'}")
    print(f"Filter     : status={args.status}"
          + (f"  only={args.only!r}" if args.only else ""))
    print(f"Queued     : {len(queue)}"
          + (f"   (skipping {len(skipped)} already in sent_log.json)" if skipped else ""))
    if args.test_to:
        print(f"TEST MODE  : every message redirected to {args.test_to}")
    print()

    if not queue:
        print("Nothing to do.")
        return

    # preview
    if args.preview:
        for r in queue[: args.preview]:
            subject = render(subject_tpl, r, extra)
            text = render(text_tpl, r, extra)
            check_unfilled(subject, text, None, r.university)
            print("─" * 72)
            print(f"To      : {r.email}   ({r.university}, {r.city})")
            print(f"Subject : {subject}")
            print()
            print(text)
        print("─" * 72)
        print(f"\n[preview] {len(queue)} recipient(s) queued. Nothing was sent.")
        return

    # Render everything up front so a template mistake fails before the first
    # message goes out, not halfway through the list.
    rendered = []
    for r in queue:
        subject = render(subject_tpl, r, extra)
        text = render(text_tpl, r, extra)
        html = render(html_tpl, r, extra) if html_tpl else None
        check_unfilled(subject, text, html, r.university)
        rendered.append((r, subject, text, html))

    if not (args.send and args.confirm):
        for r, subject, _, _ in rendered:
            print(f"  would send → {r.email:<45} {subject}")
        print()
        if args.send and not args.confirm:
            print("--send was given without --confirm, so nothing was sent.")
        print(f"[dry run] {len(rendered)} email(s) NOT sent. "
              f"Add --send --confirm to transmit.")
        return

    # ── real send ────────────────────────────────────────────────────────────
    cfg.require()
    print(f"Sending {len(rendered)} email(s), ~{args.delay:.0f}s apart. Ctrl-C to stop "
          f"(progress is saved after every message).\n")

    context = ssl.create_default_context()
    sent = failed = 0
    consecutive_failures = 0

    try:
        if cfg.use_ssl:
            server = smtplib.SMTP_SSL(cfg.host, cfg.port, context=context, timeout=30)
        else:
            server = smtplib.SMTP(cfg.host, cfg.port, timeout=30)
            server.ehlo()
            server.starttls(context=context)
            server.ehlo()
        server.login(cfg.user, cfg.password)
    except (smtplib.SMTPException, OSError) as e:
        sys.exit(f"Could not connect or authenticate to {cfg.host}:{cfg.port} — {e}")

    try:
        for i, (r, subject, text, html) in enumerate(rendered, 1):
            msg = build_message(cfg, r, subject, text, html, args.test_to)
            target = args.test_to or r.email
            try:
                server.send_message(msg)
            except smtplib.SMTPException as e:
                failed += 1
                consecutive_failures += 1
                print(f"  [{i}/{len(rendered)}] FAILED {target} — {e}")
                # A run of failures usually means throttling or a blocked
                # account; hammering on would only make that worse.
                if consecutive_failures >= 5:
                    print("\nAborting: 5 consecutive failures.")
                    break
                continue

            consecutive_failures = 0
            sent += 1
            print(f"  [{i}/{len(rendered)}] sent → {target}   ({r.university})")

            if not args.test_to:
                log[r.key] = {
                    "university": r.university,
                    "email": r.email,
                    "subject": subject,
                    "sent_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
                }
                save_log(log)

            if i < len(rendered):
                time.sleep(max(0.0, args.delay + random.uniform(0, args.delay * 0.35)))
    except KeyboardInterrupt:
        print("\nStopped by user.")
    finally:
        try:
            server.quit()
        except smtplib.SMTPException:
            pass

    print(f"\nDone. sent={sent}  failed={failed}  remaining={len(rendered) - sent - failed}")
    if not args.test_to and sent:
        print(f"Recorded in {SENT_LOG.relative_to(REPO)} — re-running skips these.")


if __name__ == "__main__":
    main()
